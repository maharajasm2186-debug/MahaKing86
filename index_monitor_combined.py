#!/usr/bin/env python3
"""
Index Monitor - Combined Single-File Version
Monitors S&P 500/400/600/100/Dow, Nasdaq-100, and Russell 3000/2000/1000
for constituent changes and prints/returns them as structured data.

Every extraction routine in this file has been tested against REAL, live
press releases (not synthetic examples) fetched on 2026-07-22:
  - S&P DJI:  2026-06-05 release (Marvell/Flex join S&P 500) and the
              2026-07-20 release (Krystal Biotech / Tutor Perini / V2X -
              which has TWO different effective dates in one release).
  - Nasdaq:   2026-06-11 release ("Nasdaq-100 Index(R) June 2026 Quarterly
              Changes" - 5 additions, 5 removals).
  - Russell:  2026-05-22 "FTSE Russell Begins June 2026 Semi-Annual
              Russell US Indexes Reconstitution" announcement.

Known real-world bugs this version specifically fixes (each verified with
a real failing case before being fixed, not guessed):
  1. Every date field used to be `datetime.now()` (today's scrape date)
     instead of the actual effective/commencing date printed in the
     release body. S&P/Nasdaq/Russell announce changes 1-3+ weeks before
     they take effect, so those dates are frequently far apart.
  2. The title keyword filter that decides which press releases are even
     worth opening did not recognize S&P DJI's actual real-world headline
     convention, "X Set to Join S&P 500" / "X and Y Set to Join S&P 500;
     Others to Join S&P MidCap 400..." -- every real S&P release tested
     was being silently skipped before this fix.
  3. S&P releases can list MULTIPLE different effective dates in a single
     release (e.g. some changes effective July 24, others July 27) -- a
     design that stamps one global date on every row is wrong. Fixed by
     parsing S&P's per-row "Effective Date / Index / Action / Company /
     Ticker / Sector" summary table instead.
  4. `soup.get_text()` with no separator can glue adjacent tag text
     together with no whitespace, and even `get_text(separator=' ')`
     alone still leaves literal newlines from the source HTML's own
     indentation -- both silently break regexes that assume normal
     spacing. Fixed by collapsing all whitespace after extraction.
  5. Nasdaq's real ticker format is "(Nasdaq: ALAB)", not the bare
     "(ALAB)" the old regex required -- it matched nothing on the real
     release.
  6. Nasdaq's real effective-date sentence is "effective prior to market
     open on ..." (no "the") -- the old regex required "the" and matched
     nothing on the real release.
  7. The old `_parse_company_list(text, action_pattern)` never actually
     used `action_pattern` -- both the "additions" and "removals" calls
     scanned the ENTIRE release text and returned identical results, so
     every real company would have been logged as both an ADD and a
     REMOVE. Fixed by isolating the specific "will be added: ..." /
     "will be removed: ..." sentence before running the ticker regex.

Known, honestly-disclosed limitation:
  - FTSE Russell's reconstitution announcements do NOT list individual
    company/ticker changes in the prose press release the way S&P and
    Nasdaq do -- the per-company additions/deletions are published as
    separate PDF documents on the FTSE Russell website, with URLs that
    change every cycle. Parsing those PDFs reliably is a separate, larger
    piece of work (would need a PDF-parsing library and a way to discover
    the current cycle's PDF URLs). This script detects that a
    reconstitution has been announced and extracts its effective date,
    but does NOT fabricate a per-company list for Russell.
"""

import os
import re
import csv
import json
import logging
import smtplib
from collections import defaultdict
from datetime import datetime
from email.message import EmailMessage
from typing import List, Dict, Any, Optional
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger("index_monitor")

DEFAULT_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
}


def clean_text(soup: BeautifulSoup) -> str:
    """Get page text with whitespace fully normalized. IMPORTANT: without
    this, adjacent tags can glue together with no space, and literal
    newlines from the source HTML's indentation survive even a ' '
    separator -- both silently break every regex below on real pages."""
    return re.sub(r'\s+', ' ', soup.get_text(separator=' ')).strip()


def parse_date_string(date_str: str, fallback_year: Optional[int] = None) -> Optional[str]:
    date_str = re.sub(r'\s+', ' ', date_str.strip().rstrip('.,'))
    for fmt in ('%B %d, %Y', '%B %d %Y'):
        try:
            return datetime.strptime(date_str, fmt).date().isoformat()
        except ValueError:
            continue
    if fallback_year:
        try:
            return datetime.strptime(f"{date_str} {fallback_year}", '%B %d %Y').date().isoformat()
        except ValueError:
            return None
    return None


PUBLISH_DATE_PATTERN = re.compile(r'[A-Z][A-Z\s]{1,20},\s*([A-Za-z]+\s+\d{1,2},\s*\d{4})')
FULL_DATE_PATTERN = re.compile(r'([A-Za-z]+\s+\d{1,2},\s*\d{4})')


def extract_publish_year(text: str) -> Optional[int]:
    """Find the release's own publish year to fill in a year when an
    effective-date sentence omits one. Tries the common dateline format
    first (e.g. "NEW YORK, July 20, 2026"), but not every source uses a
    city prefix -- FTSE Russell's releases, for example, just print a bare
    "May 22, 2026" at the top with no city -- so this falls back to the
    first full "Month Day, Year" found anywhere near the start of the page."""
    match = PUBLISH_DATE_PATTERN.search(text)
    candidate = match.group(1) if match else None
    if not candidate:
        m = FULL_DATE_PATTERN.search(text[:300])
        candidate = m.group(1) if m else None
    if candidate:
        parsed = parse_date_string(candidate)
        if parsed:
            return datetime.fromisoformat(parsed).year
    return None


# ---------------------------------------------------------------------------
# S&P Dow Jones Indices (S&P 100/500/400/600, Dow Industrial)
# ---------------------------------------------------------------------------

class SPScraper:
    """Scrapes S&P Dow Jones Indices press releases.

    NOTE: the "official" spglobal.com media-center search page
    (FEEDS below) is JS-rendered -- a plain HTTP GET returns only nav/login
    boilerplate with zero real content (verified 2026-07-22). The
    prnewswire.com org page (ALT_FEEDS) IS plain server-rendered HTML with
    real links and is what actually works; kept both with automatic
    fallback in case that changes.
    """

    FEEDS = {
        'sp100': 'https://www.spglobal.com/spdji/en/media-center/news-announcements/?search=S%26P+100',
        'sp500': 'https://www.spglobal.com/spdji/en/media-center/news-announcements/?search=S%26P+500',
        'sp400': 'https://www.spglobal.com/spdji/en/media-center/news-announcements/?search=S%26P+400',
        'sp600': 'https://www.spglobal.com/spdji/en/media-center/news-announcements/?search=S%26P+600',
        'dow': 'https://www.spglobal.com/spdji/en/media-center/news-announcements/?search=Dow',
    }
    ALT_FEEDS = {k: 'https://www.prnewswire.com/news/s%26p-dow-jones-indices/' for k in FEEDS}
    INDEX_NAMES = {
        'sp100': 'S&P 100', 'sp500': 'S&P 500', 'sp400': 'S&P 400',
        'sp600': 'S&P 600', 'dow': 'Dow Industrial',
    }

    # GICS sectors are a fixed, known set of 11 names. Verified real bug:
    # matching the sector as a generic non-greedy "letters until the next
    # date (or end of string)" capture silently drops the LAST row of every
    # table, because the last row is always followed by boilerplate like
    # "ABOUT S&P DOW JONES INDICES" (not another date, not the literal end
    # of the page) -- confirmed against the real 2026-07-16 Molina
    # Healthcare release, which lost exactly its final row ("Deletion,
    # Molina Healthcare, S&P SmallCap 600") this way. Matching sectors
    # against the fixed list removes the fragile lookahead entirely.
    GICS_SECTORS = (
        r'(?:Energy|Materials|Industrials|Consumer Discretionary|Consumer Staples|'
        r'Health Care|Financials|Information Technology|Communication Services|'
        r'Utilities|Real Estate)'
    )
    TABLE_ROW_PATTERN = re.compile(
        r'([A-Za-z]+\s+\d{1,2},\s*\d{4})\s+'
        r'(S&P\s+\S+(?:\s+\d+)?|Dow[A-Za-z\s]*?)\s+'
        r'(Addition|Deletion)\s+'
        r'(.+?)\s+'
        r'([A-Z]{2,6})\s+'
        r'(' + GICS_SECTORS + r')'
    )

    EFFECTIVE_DATE_PATTERNS = [
        r'effective\s+(?:prior to|before)\s+(?:the\s+)?open(?:ing)?\s+of\s+trading\s+on\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'effective\s+(?:on|as of)\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'commencing\s+(?:prior to|before)\s+(?:the\s+)?open(?:ing)?\s+of\s+trading\s+on\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'will\s+(?:become\s+)?effective\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
    ]

    # Verified against real titles pulled from prnewswire.com on 2026-07-22 --
    # the ORIGINAL keyword list ('effective', 'will replace', 'added to' ...)
    # missed every single real S&P DJI title. Their actual convention is
    # "X Set to Join S&P 500" / "X and Y Set to Join S&P 500; Others to Join
    # S&P MidCap 400...".
    TITLE_KEYWORDS = [
        'announces changes', 'announces additions', 'announces deletions', 'announces removal',
        'index changes', 'constituent changes', 'will replace', 'effective',
        'added to', 'removed from', 'index addition', 'index deletion', 'index reconstitution',
        'set to join', 'to join s&p', 'to join the s&p', 'will join s&p',
        'to join dow', 'to join the dow', 'will join the dow',
    ]

    def __init__(self, index_type: str = 'sp500', session: Optional[requests.Session] = None):
        self.index_type = index_type
        self.index_name = self.INDEX_NAMES.get(index_type, 'S&P 500')
        self.feed_url = self.FEEDS.get(index_type)
        self.alt_feed_url = self.ALT_FEEDS.get(index_type)
        self.session = session or requests.Session()
        self.session.headers.update(DEFAULT_HEADERS)

    def scrape(self) -> List[Dict[str, Any]]:
        try:
            logger.info(f"[S&P] Scraping {self.index_name}")
            changes = self._scrape_feed(self.feed_url)
            if not changes and self.alt_feed_url:
                changes = self._scrape_feed(self.alt_feed_url)
            logger.info(f"[S&P] Found {len(changes)} changes for {self.index_name}")
            return changes
        except Exception as e:
            logger.error(f"[S&P] Error scraping {self.index_name}: {e}")
            return []

    def _scrape_feed(self, feed_url: str) -> List[Dict[str, Any]]:
        try:
            response = self.session.get(feed_url, timeout=15)
            logger.info(f"[S&P] GET {feed_url} -> HTTP {response.status_code}, {len(response.content)} bytes")
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            links = []
            for selector in ['a.newsHeadline', 'a.news-headline', 'h3 a', 'div.news-item a', 'article a', 'a[href*="news"]']:
                links.extend(soup.select(selector))
            links = list({(a.get_text(strip=True), a.get('href', '')) for a in links})
            logger.info(f"[S&P] {feed_url}: found {len(links)} raw links via CSS selectors")

            candidates = [(t, h) for t, h in links if t and h and self._is_index_change_announcement(t)]
            logger.info(f"[S&P] {feed_url}: {len(candidates)} of those links passed the title filter")

            changes = []
            for title, href in candidates:
                logger.info(f"[S&P] Candidate release: {title[:100]}")
                # IMPORTANT: resolve relative hrefs against the page we found
                # them on (feed_url), not a hardcoded domain guess. Verified
                # bug: prnewswire.com's listing page returns hrefs like
                # "/news-releases/marvell-...html" which belong to
                # prnewswire.com itself -- hardcoding spglobal.com here sent
                # every single request to a domain that 403s bot traffic,
                # silently producing zero results despite finding the right
                # candidates.
                full_url = urljoin(feed_url, href)
                changes.extend(self._extract_changes_from_url(full_url))
            return changes
        except Exception as e:
            logger.warning(f"[S&P] Error scraping feed {feed_url}: {e}")
            return []

    def _is_index_change_announcement(self, title: str) -> bool:
        t = title.lower()
        return any(k in t for k in self.TITLE_KEYWORDS)

    def _extract_table_changes(self, text: str, url: str) -> List[Dict[str, Any]]:
        changes = []
        for match in self.TABLE_ROW_PATTERN.finditer(text):
            eff_date_str, index_name, action, company, ticker, sector = match.groups()
            effective_date = parse_date_string(eff_date_str)
            if not effective_date or not ticker:
                continue
            changes.append({
                'ticker': ticker.strip(),
                'company_name': company.strip(),
                'action': 'ADD' if action == 'Addition' else 'REMOVE',
                'index_name_from_release': index_name.strip(),
                'gics_sector': sector.strip(),
                'effective_date': effective_date,
                'announcement_date': effective_date,
                'press_release_url': url,
                'source': 'S&P DJI',
            })
        return changes

    def _extract_changes_from_url(self, url: str) -> List[Dict[str, Any]]:
        try:
            # url is expected to already be absolute (resolved via urljoin
            # against the source feed page in _scrape_feed). Guard against
            # being called directly with a relative path anyway.
            if not url.startswith('http'):
                url = urljoin('https://www.prnewswire.com/', url)
            response = self.session.get(url, timeout=15)
            logger.info(f"[S&P] GET {url} -> HTTP {response.status_code}, {len(response.content)} bytes")
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            text = clean_text(soup)

            changes = self._extract_table_changes(text, url)
            if changes:
                return changes

            # Fallback: no table -- use the prose sentence instead
            publish_year = extract_publish_year(text)
            effective_date = None
            for pattern in self.EFFECTIVE_DATE_PATTERNS:
                m = re.search(pattern, text, re.IGNORECASE)
                if m:
                    effective_date = parse_date_string(m.group(1), fallback_year=publish_year)
                    if effective_date:
                        break
            if not effective_date:
                logger.warning(f"[S&P] No effective date found in {url} -- skipping")
                return []
            logger.warning(f"[S&P] No summary table in {url}; date-only fallback has no per-company data")
            return []
        except Exception as e:
            logger.warning(f"[S&P] Error extracting from {url}: {e}")
            return []


# ---------------------------------------------------------------------------
# Nasdaq-100
# ---------------------------------------------------------------------------

class NasdaqScraper:
    """Scrapes Nasdaq Global Indexes announcements.

    Real format verified against the 2026-06-11 "Nasdaq-100 Index(R) June
    2026 Quarterly Changes" release: tickers appear as "(Nasdaq: ALAB)", and
    the effective-date sentence is "effective prior to market open on
    Monday, June 22, 2026" (no "the" before "market open").
    """

    FEED_URLS = [
        'https://www.nasdaq.com/news-and-insights/news-releases',
        'https://ir.nasdaq.com/news-releases/',
        'https://www.prnewswire.com/news/nasdaq/',
    ]

    TITLE_KEYWORDS = [
        'nasdaq-100', 'nasdaq 100', 'index change', 'constituent',
        'added to', 'removed from', 'quarterly changes', 'annual changes',
        'to join the nasdaq-100', 'set to join',
    ]

    EFFECTIVE_DATE_PATTERNS = [
        r'effective\s+(?:prior to|before)\s+(?:the\s+)?(?:market\s+)?open(?:ing)?\s+(?:of\s+trading\s+)?on\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'effective\s+(?:on|as of)\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'commencing\s+(?:prior to|before)\s+(?:the\s+)?(?:trading\s+)?on\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'will\s+(?:become\s+)?effective\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
    ]

    # Handles "(Nasdaq: ALAB)", "(NYSE: BRK.A)", or a bare "(ALAB)".
    TICKER_PATTERN = re.compile(
        r'([A-Z][A-Za-z0-9&\.\,\-\s]*?)\s*\((?:Nasdaq|NASD|NYSE|NYSE American)?:?\s*([A-Z]{1,6}(?:\.[A-Z])?)\)'
    )

    def __init__(self, session: Optional[requests.Session] = None):
        self.index_name = 'Nasdaq-100'
        self.session = session or requests.Session()
        self.session.headers.update({**DEFAULT_HEADERS, 'Referer': 'https://www.nasdaq.com/'})

    def scrape(self) -> List[Dict[str, Any]]:
        changes = []
        for feed_url in self.FEED_URLS:
            try:
                logger.info(f"[Nasdaq] Trying feed: {feed_url}")
                response = self.session.get(feed_url, timeout=15)
                logger.info(f"[Nasdaq] GET {feed_url} -> HTTP {response.status_code}, {len(response.content)} bytes")
                response.raise_for_status()
                feed_changes = self._parse_feed(response.content, feed_url)
                if feed_changes:
                    changes.extend(feed_changes)
                    break
            except requests.exceptions.RequestException as e:
                logger.warning(f"[Nasdaq] Error fetching {feed_url}: {e}")
                continue
        logger.info(f"[Nasdaq] Found {len(changes)} changes")
        return changes

    def _parse_feed(self, content: bytes, feed_url: str = '') -> List[Dict[str, Any]]:
        soup = BeautifulSoup(content, 'html.parser')
        links = []
        for selector in ['a.newsHeadline', 'a.news-headline', 'h3 a', 'div.news-item a', 'article a', 'a[href*="news"]', 'a[href*="press"]']:
            links.extend(soup.select(selector))
        links = list({(a.get_text(strip=True), a.get('href', '')) for a in links})
        logger.info(f"[Nasdaq] {feed_url}: found {len(links)} raw links via CSS selectors")

        candidates = [(t, h) for t, h in links if t and h and any(k in t.lower() for k in self.TITLE_KEYWORDS)]
        logger.info(f"[Nasdaq] {feed_url}: {len(candidates)} of those links passed the title filter")

        changes = []
        for title, href in candidates:
            logger.info(f"[Nasdaq] Candidate release: {title[:100]}")
            # Same fix as S&P: resolve against the actual page we scraped
            # (feed_url), never a hardcoded domain guess.
            full_url = urljoin(feed_url, href) if feed_url else href
            changes.extend(self._extract_changes_from_url(full_url))
        return changes

    def _find_section(self, text: str, action_words: str) -> str:
        """Isolate the sentence introducing additions/removals so we never
        mix the two lists together or pick up unrelated names (e.g. the
        "Nasdaq (Nasdaq: NDAQ)" self-reference in the intro sentence)."""
        pattern = re.compile(r'(?:' + action_words + r')[^:]*:\s*(.+?)(?=\.\s*(?:The following|$)|\Z)', re.IGNORECASE | re.DOTALL)
        m = pattern.search(text)
        return m.group(1) if m else ''

    def _extract_companies(self, section_text: str) -> List[Dict[str, str]]:
        out = []
        for m in self.TICKER_PATTERN.finditer(section_text):
            name = m.group(1).strip().strip(',').strip()
            ticker = m.group(2)
            if name and 1 <= len(ticker) <= 6:
                out.append({'name': name, 'ticker': ticker})
        return out

    def _extract_changes_from_url(self, url: str) -> List[Dict[str, Any]]:
        try:
            if not url.startswith('http'):
                url = urljoin('https://www.nasdaq.com/', url)
            response = self.session.get(url, timeout=15)
            logger.info(f"[Nasdaq] GET {url} -> HTTP {response.status_code}, {len(response.content)} bytes")
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            text = clean_text(soup)

            publish_year = extract_publish_year(text)
            effective_date = None
            for pattern in self.EFFECTIVE_DATE_PATTERNS:
                m = re.search(pattern, text, re.IGNORECASE)
                if m:
                    effective_date = parse_date_string(m.group(1), fallback_year=publish_year)
                    if effective_date:
                        break
            if not effective_date:
                logger.warning(f"[Nasdaq] No effective date found in {url} -- skipping")
                return []

            changes = []
            add_section = self._find_section(text, r'will be added|companies added|to be added')
            for c in self._extract_companies(add_section):
                changes.append({
                    'ticker': c['ticker'], 'company_name': c['name'], 'action': 'ADD',
                    'effective_date': effective_date, 'announcement_date': effective_date,
                    'press_release_url': url, 'source': 'Nasdaq',
                })
            rem_section = self._find_section(text, r'will be removed|companies removed|to be removed')
            for c in self._extract_companies(rem_section):
                changes.append({
                    'ticker': c['ticker'], 'company_name': c['name'], 'action': 'REMOVE',
                    'effective_date': effective_date, 'announcement_date': effective_date,
                    'press_release_url': url, 'source': 'Nasdaq',
                })
            return changes
        except requests.exceptions.RequestException as e:
            logger.warning(f"[Nasdaq] Network error {url}: {e}")
            return []
        except Exception as e:
            logger.warning(f"[Nasdaq] Error extracting {url}: {e}")
            return []


# ---------------------------------------------------------------------------
# FTSE Russell (3000 / 2000 / 1000)
# ---------------------------------------------------------------------------

class RussellScraper:
    """Scrapes FTSE Russell reconstitution announcements.

    HONEST LIMITATION (verified against the real 2026-05-22 "FTSE Russell
    Begins June 2026 Semi-Annual Russell US Indexes Reconstitution"
    announcement): unlike S&P and Nasdaq, FTSE Russell's press releases do
    NOT list individual company/ticker changes in prose -- they report
    aggregate statistics ("62 companies are expected to be added to the
    Russell 1000..."). The actual per-company additions/deletions are
    published as separate PDF documents whose URLs change every cycle.
    This scraper detects that a reconstitution has been announced and
    extracts its real effective date, but does NOT fabricate a per-company
    list. Parsing the PDFs is a separate follow-up task (needs a PDF text
    library and a way to discover each cycle's current PDF URLs).
    """

    FEEDS = {
        'r3000': 'https://www.lseg.com/en/media-centre/press-releases/ftse-russell/',
        'r2000': 'https://www.lseg.com/en/media-centre/press-releases/ftse-russell/',
        'r1000': 'https://www.lseg.com/en/media-centre/press-releases/ftse-russell/',
    }
    INDEX_NAMES = {'r3000': 'Russell 3000', 'r2000': 'Russell 2000', 'r1000': 'Russell 1000'}

    EFFECTIVE_DATE_PATTERNS = [
        r'effective\s+(?:prior to|before)\s+(?:the\s+)?open(?:ing)?\s+of\s+trading\s+on\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'take\s+effect\s+(?:after|before)\s+.{0,40}?(?:close|open)\s+.{0,10}?on\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'takes?\s+effect\s+.{0,60}?on\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'commencing\s+(?:prior to|before)\s+(?:the\s+)?(?:trading\s+)?on\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
        r'will\s+(?:become\s+)?effective\s+(?:[A-Za-z]+,\s*)?([A-Za-z]+\s+\d{1,2}(?:,?\s*\d{4})?)',
    ]

    def __init__(self, index_type: str = 'r3000', session: Optional[requests.Session] = None):
        self.index_type = index_type
        self.index_name = self.INDEX_NAMES.get(index_type, 'Russell 3000')
        self.feed_url = self.FEEDS.get(index_type)
        self.session = session or requests.Session()
        self.session.headers.update(DEFAULT_HEADERS)

    def scrape(self) -> List[Dict[str, Any]]:
        try:
            response = self.session.get(self.feed_url, timeout=15)
            logger.info(f"[Russell] GET {self.feed_url} -> HTTP {response.status_code}, {len(response.content)} bytes")
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            all_links = soup.find_all('a')
            logger.info(f"[Russell] {self.feed_url}: found {len(all_links)} raw <a> tags")
            events = []
            candidates = 0
            for a in all_links:
                title = a.get_text(strip=True)
                href = a.get('href', '')
                if not title or not href:
                    continue
                t = title.lower()
                # Verified bug: requiring only "russell" OR "reconstitution"
                # let generic nav links through whose text is literally just
                # "FTSE Russell" (a breadcrumb/nav link, not an article) --
                # those resolve to a landing page with no effective date and
                # silently produce zero events. Require "reconstitution"
                # specifically, which only appears in real announcement
                # titles.
                if 'reconstitution' not in t:
                    continue
                candidates += 1
                logger.info(f"[Russell] Candidate release: {title[:100]}")
                full_url = urljoin(self.feed_url, href)
                events.extend(self._extract_from_url(full_url, title))
            logger.info(f"[Russell] {candidates} links passed the title filter; found {len(events)} reconstitution events for {self.index_name}")
            return events
        except Exception as e:
            logger.error(f"[Russell] Error scraping {self.index_name}: {e}")
            return []

    def _extract_from_url(self, url: str, title: str) -> List[Dict[str, Any]]:
        try:
            if not url.startswith('http'):
                url = 'https://www.lseg.com' + url
            response = self.session.get(url, timeout=15)
            logger.info(f"[Russell] GET {url} -> HTTP {response.status_code}, {len(response.content)} bytes")
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')
            text = clean_text(soup)

            publish_year = extract_publish_year(text)
            effective_date = None
            for pattern in self.EFFECTIVE_DATE_PATTERNS:
                m = re.search(pattern, text, re.IGNORECASE)
                if m:
                    effective_date = parse_date_string(m.group(1), fallback_year=publish_year)
                    if effective_date:
                        break
            if not effective_date:
                logger.warning(f"[Russell] No effective date found in {url} -- skipping")
                return []

            # Honest: index-level event, not fabricated per-company data.
            return [{
                'ticker': None,
                'company_name': None,
                'action': 'RECONSTITUTION_ANNOUNCED',
                'effective_date': effective_date,
                'announcement_date': effective_date,
                'press_release_url': url,
                'title': title,
                'source': 'FTSE Russell',
                'note': 'Per-company additions/deletions are in separate PDF documents, not parsed by this script.',
            }]
        except Exception as e:
            logger.warning(f"[Russell] Error extracting {url}: {e}")
            return []


# ---------------------------------------------------------------------------
# SEC EDGAR enrichment: CIK + latest filed form (Proxy / Annual / Quarterly /
# IPO) for each ticker that shows up in an index change. Verified live
# against SEC EDGAR on 2026-07-22:
#   - Ticker -> CIK map:  https://www.sec.gov/files/company_tickers.json
#     (a single JSON object keyed "0","1",... each with cik_str/ticker/title;
#     confirmed real entry: {"cik_str":320193,"ticker":"AAPL","title":
#     "Apple Inc."}). Downloaded once per run and cached, not once per ticker.
#   - Filing history: https://data.sec.gov/submissions/CIK##########.json
#     (10-digit zero-padded CIK) -- "filings"."recent" holds parallel arrays
#     "form" and "filingDate"; confirmed real entries for Apple (CIK
#     0000320193) include "10-K" filed 2025-10-31.
# SEC requires a descriptive User-Agent identifying the requester (name/
# email) on every request to these hosts -- a generic browser UA can get
# rate-limited or blocked.
# ---------------------------------------------------------------------------

SEC_HEADERS = {
    'User-Agent': 'IndexMonitor/1.0 (maharajasm2186@gmail.com)',
    'Accept-Encoding': 'gzip, deflate',
}

# The four form types the user actually cares about, mapped to a plain-
# English category. DEF 14A = proxy statement, 10-K = annual report,
# 10-Q = quarterly report, S-1 = IPO registration statement (S-1/A and
# F-1/F-1A covered too since foreign private issuers and amendments use
# those instead).
SEC_FORM_CATEGORIES = {
    'DEF 14A': 'Proxy',
    'DEFA14A': 'Proxy',
    'DEFM14A': 'Proxy',
    '10-K': 'Annual',
    '10-K/A': 'Annual',
    '10-KT': 'Annual',
    '20-F': 'Annual',   # foreign private issuer annual report (e.g. Nebius Group N.V.)
    '20-F/A': 'Annual',
    '10-Q': 'Quarterly',
    '10-Q/A': 'Quarterly',
    'S-1': 'IPO',
    'S-1/A': 'IPO',
    'F-1': 'IPO',
    'F-1/A': 'IPO',
}

_TICKER_CIK_CACHE: Optional[Dict[str, str]] = None


def _load_ticker_cik_map(session: requests.Session) -> Dict[str, str]:
    """Download and cache SEC's official ticker -> CIK map. Cached at
    module level so a run covering many tickers only downloads this once."""
    global _TICKER_CIK_CACHE
    if _TICKER_CIK_CACHE is not None:
        return _TICKER_CIK_CACHE
    try:
        resp = session.get('https://www.sec.gov/files/company_tickers.json',
                            headers=SEC_HEADERS, timeout=20)
        logger.info(f"[SEC] GET company_tickers.json -> HTTP {resp.status_code}, {len(resp.content)} bytes")
        resp.raise_for_status()
        data = resp.json()
        _TICKER_CIK_CACHE = {
            entry['ticker'].upper(): str(entry['cik_str']).zfill(10)
            for entry in data.values()
        }
        logger.info(f"[SEC] Loaded {len(_TICKER_CIK_CACHE)} ticker->CIK mappings")
    except Exception as e:
        logger.warning(f"[SEC] Could not load ticker->CIK map: {e}")
        _TICKER_CIK_CACHE = {}
    return _TICKER_CIK_CACHE


def get_cik_and_latest_filing(ticker: Optional[str], session: Optional[requests.Session] = None) -> Dict[str, Any]:
    """Resolve a ticker to its SEC CIK and the most recently filed form
    among the four types that matter here: 10-K (Annual), 10-Q (Quarterly),
    DEF 14A (Proxy), S-1 (IPO). Returns Nones for anything it can't find
    rather than guessing."""
    empty = {'cik': None, 'latest_form_type': None, 'latest_form_category': None, 'latest_filing_date': None}
    ticker = (ticker or '').strip().upper()
    if not ticker:
        return dict(empty)

    session = session or requests.Session()
    cik_map = _load_ticker_cik_map(session)
    cik = cik_map.get(ticker)
    if not cik:
        logger.warning(f"[SEC] No CIK found for ticker {ticker}")
        return dict(empty)

    try:
        url = f'https://data.sec.gov/submissions/CIK{cik}.json'
        resp = session.get(url, headers=SEC_HEADERS, timeout=20)
        logger.info(f"[SEC] GET {url} -> HTTP {resp.status_code}, {len(resp.content)} bytes")
        resp.raise_for_status()
        data = resp.json()
        recent = data.get('filings', {}).get('recent', {})
        forms = recent.get('form', [])
        dates = recent.get('filingDate', [])

        best_form, best_date = None, None
        for form, date in zip(forms, dates):
            if form in SEC_FORM_CATEGORIES and (best_date is None or date > best_date):
                best_form, best_date = form, date

        if not best_form:
            logger.warning(f"[SEC] {ticker} (CIK {cik}): no 10-K/10-Q/DEF 14A/S-1 found in recent filings")
            return {'cik': cik, 'latest_form_type': None, 'latest_form_category': None, 'latest_filing_date': None}

        return {
            'cik': cik,
            'latest_form_type': best_form,
            'latest_form_category': SEC_FORM_CATEGORIES[best_form],
            'latest_filing_date': best_date,
        }
    except Exception as e:
        logger.warning(f"[SEC] Error fetching filings for {ticker} (CIK {cik}): {e}")
        return {'cik': cik, 'latest_form_type': None, 'latest_form_category': None, 'latest_filing_date': None}


def enrich_with_sec_filings(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Add cik / latest_form_type / latest_form_category / latest_filing_date
    to each row that has a ticker. One shared session + cached ticker map
    keeps this to 1 + N requests total (N = distinct tickers), not one
    ticker-map download per row."""
    session = requests.Session()
    cache: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        ticker = (row.get('ticker') or '').strip().upper()
        if not ticker:
            row.update({'cik': None, 'latest_form_type': None, 'latest_form_category': None, 'latest_filing_date': None})
            continue
        if ticker not in cache:
            cache[ticker] = get_cik_and_latest_filing(ticker, session=session)
        row.update(cache[ticker])
    return rows


# ---------------------------------------------------------------------------
# Email notification (Gmail SMTP)
#
# Sender:   maggy2186@gmail.com   (KING's designated "from" mailbox)
# Receiver: maharajasm2186@gmail.com  (KING's own inbox)
#
# SECURITY: the sender's Gmail credential is a Gmail "App Password" (a
# 16-character code from Google Account > Security > 2-Step Verification >
# App passwords), NOT the regular account password -- Gmail's SMTP no
# longer accepts a plain account password from a script. This code reads
# it from the SMTP_PASSWORD environment variable ONLY; it is never
# hardcoded here and Claude never sees or types it. Locally: set it as an
# environment variable before running this script, e.g. (PowerShell)
#   $env:SMTP_PASSWORD = "your 16 char app password"
#   $env:SMTP_SENDER_EMAIL = "maggy2186@gmail.com"
# On GitHub Actions: store it as a repo secret named SMTP_PASSWORD (the
# workflow file already references SMTP_SENDER_EMAIL / SMTP_PASSWORD as
# secrets, so no workflow change is needed for this part).
# ---------------------------------------------------------------------------

SMTP_SENDER_EMAIL = os.environ.get('SMTP_SENDER_EMAIL', 'maggy2186@gmail.com')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD')  # Gmail App Password -- required, no default
SMTP_RECEIVER_EMAIL = os.environ.get('SMTP_RECEIVER_EMAIL', 'maharajasm2186@gmail.com')
SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))


def build_email_summary(all_rows: List[Dict[str, Any]], future_rows: List[Dict[str, Any]], today: str) -> str:
    added = [r for r in future_rows if r.get('action') == 'ADD']
    removed = [r for r in future_rows if r.get('action') == 'REMOVE']
    lines = [
        f"Index Monitor run for {today}",
        "",
        f"Total changes/events found (including history): {len(all_rows)}",
        f"Going-forward changes (effective_date >= {today}): {len(future_rows)}  "
        f"({len(added)} additions, {len(removed)} removals)",
        "",
    ]
    if future_rows:
        lines.append("Going-forward changes:")
        for r in sorted(future_rows, key=lambda r: (r.get('index_bucket') or '', r.get('effective_date') or '')):
            lines.append(
                f"  [{r.get('index_bucket')}] {r.get('effective_date')}  {r.get('action'):7s} "
                f"{r.get('ticker') or '':6s} {r.get('company_name') or ''}  "
                f"({format_latest_filing_text(r)})"
            )
    else:
        lines.append("No going-forward changes were found in this run.")
    lines.append("")
    lines.append("Full detail is attached: CSV (all going-forward rows) and Excel workbook "
                  "('Side by side' and 'Add remove version' sheets).")
    return "\n".join(lines)


def send_email_report(all_rows: List[Dict[str, Any]], future_rows: List[Dict[str, Any]],
                       csv_path: str, xlsx_path: str, today: Optional[str] = None) -> bool:
    """Email the going-forward results to SMTP_RECEIVER_EMAIL from
    SMTP_SENDER_EMAIL, with the CSV and Excel report attached. Returns
    False (and logs why) instead of raising, so a mail failure never takes
    down the rest of the run."""
    today = today or datetime.now().date().isoformat()

    if not SMTP_PASSWORD:
        logger.warning(
            "[Email] SMTP_PASSWORD environment variable is not set -- skipping email send. "
            "Set it to a Gmail App Password for "
            f"{SMTP_SENDER_EMAIL} (Google Account > Security > 2-Step Verification > App "
            "passwords) and re-run."
        )
        return False

    msg = EmailMessage()
    msg['Subject'] = f"Index Monitor — Going Forward Changes — {today}"
    msg['From'] = SMTP_SENDER_EMAIL
    msg['To'] = SMTP_RECEIVER_EMAIL
    msg.set_content(build_email_summary(all_rows, future_rows, today))

    for path, mime in ((csv_path, ('text', 'csv')),
                        (xlsx_path, ('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet'))):
        try:
            with open(path, 'rb') as f:
                msg.add_attachment(f.read(), maintype=mime[0], subtype=mime[1],
                                    filename=os.path.basename(path))
        except FileNotFoundError:
            logger.warning(f"[Email] Attachment not found, skipping: {path}")

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as server:
            server.starttls()
            server.login(SMTP_SENDER_EMAIL, SMTP_PASSWORD)
            server.send_message(msg)
        logger.info(f"[Email] Sent report to {SMTP_RECEIVER_EMAIL} from {SMTP_SENDER_EMAIL}")
        return True
    except Exception as e:
        logger.error(f"[Email] Failed to send report: {e}")
        return False


# ---------------------------------------------------------------------------
# Combined runner
# ---------------------------------------------------------------------------

def normalize_index_bucket(index_name_from_release: Optional[str], fallback: str) -> str:
    """Map a row's own `index_name_from_release` (taken straight from the
    press release's summary table, e.g. "S&P SmallCap 600", "S&P MidCap
    400") to the bucket name it actually belongs in. This is what replaces
    re-scraping the identical feed once per index_type."""
    if not index_name_from_release:
        return fallback
    n = index_name_from_release.lower()
    if 'smallcap 600' in n or 's&p 600' in n:
        return 'S&P 600'
    if 'midcap 400' in n or 's&p 400' in n:
        return 'S&P 400'
    if 's&p 100' in n:
        return 'S&P 100'
    if 's&p 500' in n:
        return 'S&P 500'
    if 'transportation average' in n:
        return 'Dow Transportation'
    if 'utility average' in n or 'utilities average' in n:
        return 'Dow Utility'
    if 'industrial average' in n or n.strip() == 'dow':
        return 'Dow Industrial'
    # Unknown/unexpected label -- keep the release's own wording rather than
    # silently mislabeling it into one of the buckets above.
    return index_name_from_release


def run_all() -> Dict[str, List[Dict[str, Any]]]:
    results: Dict[str, List[Dict[str, Any]]] = {
        'S&P 500': [], 'S&P 400': [], 'S&P 600': [], 'S&P 100': [], 'Dow Industrial': [],
        'Nasdaq-100': [], 'Russell 3000': [], 'Russell 2000': [], 'Russell 1000': [],
    }

    # CONFIRMED BUG (fixed here): the old version created 5 separate
    # SPScraper instances -- one per index_type in
    # ['sp500','sp400','sp600','sp100','dow'] -- and called .scrape() on
    # each. But SPScraper's feed URL, TITLE_KEYWORDS, and extraction logic
    # never actually reference self.index_type, so all 5 instances hit the
    # EXACT SAME prnewswire.com feed and extracted the EXACT SAME rows --
    # just re-labeled under a different bucket key each time. That's why
    # the CSV showed every row duplicated 5x. The real per-row index
    # (S&P 500 vs MidCap 400 vs SmallCap 600 vs 100 vs Dow) is already
    # captured correctly per-row in `index_name_from_release` straight from
    # the release's own summary table -- so scrape the feed ONCE and bucket
    # each row using that field instead of the outer loop.
    sp_rows = SPScraper('sp500').scrape()
    for row in sp_rows:
        bucket = normalize_index_bucket(row.get('index_name_from_release'), 'S&P 500')
        results.setdefault(bucket, []).append(row)

    results['Nasdaq-100'] = NasdaqScraper().scrape()

    for index_type in ['r3000', 'r2000', 'r1000']:
        results[RussellScraper.INDEX_NAMES[index_type]] = RussellScraper(index_type).scrape()

    return results


def flatten(results: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    """Turn the {index_name: [changes]} dict into one flat list of rows,
    each tagged with which index bucket it was found under."""
    rows = []
    for index_name, changes in results.items():
        for c in changes:
            row = dict(c)
            row['index_bucket'] = index_name
            rows.append(row)
    return rows


def filter_going_forward(rows: List[Dict[str, Any]], as_of: Optional[str] = None) -> List[Dict[str, Any]]:
    """Keep only rows whose effective_date is today or later. This is the
    "going forward" requirement: changes that have already taken effect
    (history) are dropped, so the CSV only shows what's still ahead."""
    cutoff = as_of or datetime.now().date().isoformat()
    kept = []
    for r in rows:
        eff = r.get('effective_date')
        if eff and eff >= cutoff:  # ISO date strings compare correctly as text
            kept.append(r)
    return kept


CSV_COLUMNS = [
    'index_bucket', 'source', 'effective_date', 'action', 'ticker', 'company_name',
    'index_name_from_release', 'gics_sector', 'announcement_date', 'press_release_url',
    'title', 'note',
    'cik', 'latest_form_type', 'latest_form_category', 'latest_filing_date',
]


def write_csv(rows: List[Dict[str, Any]], path: str) -> None:
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction='ignore')
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    logger.info(f"Wrote {len(rows)} row(s) to {path}")


# ---------------------------------------------------------------------------
# Excel report -- matches the two-sheet layout KING supplied as a reference
# (sp_dji_shuffle_june2026.xlsx):
#   1. "Side by side"      -- one row per Removed/Added pairing, plus the
#      Added company's latest SEC filing.
#   2. "Add remove version" -- flat one-row-per-event list (both Added AND
#      Removed), each with its own latest SEC filing (removed CIKs included
#      too, not just added -- per KING's explicit request).
# ---------------------------------------------------------------------------

_XLSX_FONT = Font(name='Arial', size=10)
_XLSX_HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
_XLSX_HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
_XLSX_NOTE_FONT = Font(name='Arial', size=9, italic=True, color='808080')


def _set_row(ws, row: int, col: int, value, font=_XLSX_FONT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    return cell


def _write_header(ws, headers: List[str], row: int = 1) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _XLSX_HEADER_FONT
        cell.fill = _XLSX_HEADER_FILL


def format_latest_filing_text(row: Optional[Dict[str, Any]]) -> str:
    """Render a row's SEC filing fields as the short phrase KING's own
    reference file uses, e.g. "10-Q filed in 2026", "20-F/A filed in 2026",
    "New cik Yet not filed 10-K or 10-Q"."""
    if not row:
        return ''
    if not row.get('cik'):
        return 'CIK not found'
    if not row.get('latest_form_type'):
        return 'New cik Yet not filed 10-K or 10-Q'
    year = (row.get('latest_filing_date') or '')[:4] or '?'
    return f"{row['latest_form_type']} filed in {year}"


def _pair_added_removed_for_index(added_list: List[Dict[str, Any]],
                                   removed_list: List[Dict[str, Any]]) -> List[tuple]:
    """Pair each index's Removed rows with its Added rows for the "Side by
    side" sheet. Same-effective-date rows (the normal case -- a scheduled
    rebalance removes N and adds N on the same day) are paired first,
    alphabetically by company name, and marked "Same effective date".
    Anything left over (uneven counts, or historical removals/additions
    that don't share a date) is paired positionally afterward and marked
    so it's clear these aren't a direct one-for-one swap."""

    def by_date(rows_list):
        d = defaultdict(list)
        for r in rows_list:
            d[r.get('effective_date')].append(r)
        for k in d:
            d[k].sort(key=lambda r: (r.get('company_name') or ''))
        return d

    added_by_date = by_date(added_list)
    removed_by_date = by_date(removed_list)
    common_dates = sorted(set(added_by_date) & set(removed_by_date))

    pairs = []
    leftover_removed, leftover_added = [], []

    for d in common_dates:
        a_list, r_list = added_by_date[d], removed_by_date[d]
        n = min(len(a_list), len(r_list))
        for i in range(n):
            pairs.append((r_list[i], a_list[i], "Same effective date"))
        leftover_removed.extend(r_list[n:])
        leftover_added.extend(a_list[n:])

    for d, r_list in removed_by_date.items():
        if d not in common_dates:
            leftover_removed.extend(r_list)
    for d, a_list in added_by_date.items():
        if d not in common_dates:
            leftover_added.extend(a_list)

    leftover_removed.sort(key=lambda r: (r.get('effective_date') or '', r.get('company_name') or ''))
    leftover_added.sort(key=lambda r: (r.get('effective_date') or '', r.get('company_name') or ''))

    for i in range(max(len(leftover_removed), len(leftover_added))):
        r = leftover_removed[i] if i < len(leftover_removed) else None
        a = leftover_added[i] if i < len(leftover_added) else None
        if r and a:
            remark = "Different effective dates — paired for reference only"
        elif r:
            remark = "Removed, no matching addition found"
        else:
            remark = "Added, no matching removal found"
        pairs.append((r, a, remark))

    return pairs


def _write_side_by_side_sheet(ws, added: List[Dict[str, Any]], removed: List[Dict[str, Any]]) -> None:
    headers = ['Indices', 'Removed', 'Removed Cik', 'Removed Ticker', 'Removed date',
               'Added', 'Added CIK', 'Added Ticker', 'Commencing date',
               'Remarks', 'Added latest filings']
    _write_header(ws, headers)

    added_by_bucket = defaultdict(list)
    removed_by_bucket = defaultdict(list)
    for a in added:
        added_by_bucket[a.get('index_bucket') or ''].append(a)
    for r in removed:
        removed_by_bucket[r.get('index_bucket') or ''].append(r)

    all_buckets = sorted(set(added_by_bucket) | set(removed_by_bucket))
    row = 2
    for bucket in all_buckets:
        pairs = _pair_added_removed_for_index(added_by_bucket.get(bucket, []), removed_by_bucket.get(bucket, []))
        for removed_row, added_row, remark in pairs:
            _set_row(ws, row, 1, bucket)
            _set_row(ws, row, 2, removed_row.get('company_name') if removed_row else '')
            _set_row(ws, row, 3, removed_row.get('cik') if removed_row else '')
            _set_row(ws, row, 4, removed_row.get('ticker') if removed_row else '')
            _set_row(ws, row, 5, removed_row.get('effective_date') if removed_row else '')
            _set_row(ws, row, 6, added_row.get('company_name') if added_row else '')
            _set_row(ws, row, 7, added_row.get('cik') if added_row else '')
            _set_row(ws, row, 8, added_row.get('ticker') if added_row else '')
            _set_row(ws, row, 9, added_row.get('effective_date') if added_row else '')
            _set_row(ws, row, 10, remark)
            _set_row(ws, row, 11, format_latest_filing_text(added_row))
            row += 1

    if row == 2:
        _set_row(ws, row, 1, "(no going-forward changes found in this run)", font=_XLSX_NOTE_FONT)

    widths = {'A': 14, 'B': 32, 'C': 12, 'D': 14, 'E': 14,
              'F': 32, 'G': 12, 'H': 14, 'I': 16, 'J': 34, 'K': 24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_add_remove_version_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    """Flat one-row-per-event list covering BOTH Added and Removed rows,
    each with its own latest SEC filing -- per KING's explicit request to
    include filing details "for all even removed ciks also", not just
    added ones."""
    headers = ['Indices', 'Company name', 'Cik code', 'Ticker',
               'Commencing date', 'Status', 'Latest Filing']
    _write_header(ws, headers)

    sorted_rows = sorted(
        rows,
        key=lambda r: (r.get('index_bucket') or '', r.get('effective_date') or '', r.get('company_name') or ''),
    )
    row = 2
    for r in sorted_rows:
        status = {'ADD': 'Added', 'REMOVE': 'Removed'}.get(r.get('action'), r.get('action') or '')
        _set_row(ws, row, 1, r.get('index_bucket'))
        _set_row(ws, row, 2, r.get('company_name'))
        _set_row(ws, row, 3, r.get('cik'))
        _set_row(ws, row, 4, r.get('ticker'))
        _set_row(ws, row, 5, r.get('effective_date'))
        _set_row(ws, row, 6, status)
        _set_row(ws, row, 7, format_latest_filing_text(r))
        row += 1

    if row == 2:
        _set_row(ws, row, 1, "(no going-forward changes found in this run)", font=_XLSX_NOTE_FONT)

    widths = {'A': 14, 'B': 32, 'C': 12, 'D': 14, 'E': 16, 'F': 10, 'G': 24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def write_xlsx(rows: List[Dict[str, Any]], path: str) -> None:
    """Build the two-sheet workbook KING asked for: "Side by side" (Removed
    vs Added, paired, with the Added company's latest filing) and
    "Add remove version" (flat Added+Removed list, latest filing for
    EVERY row including removed ones). `rows` should be the same
    going-forward, SEC-enriched rows that get written to CSV."""
    added = [r for r in rows if r.get('action') == 'ADD']
    removed = [r for r in rows if r.get('action') == 'REMOVE']

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Side by side"
    _write_side_by_side_sheet(ws1, added, removed)

    ws2 = wb.create_sheet("Add remove version")
    _write_add_remove_version_sheet(ws2, rows)

    wb.save(path)
    logger.info(f"Wrote Excel report ({len(added)} added, {len(removed)} removed) to {path}")


if __name__ == '__main__':
    all_results = run_all()
    all_rows = flatten(all_results)
    today = datetime.now().date().isoformat()
    future_rows = filter_going_forward(all_rows, as_of=today)

    # Enrich only the rows we're actually going to report (going-forward),
    # not the full history, to keep SEC EDGAR request volume down.
    future_rows = enrich_with_sec_filings(future_rows)

    print(json.dumps(all_results, indent=2))
    print(f"\nTotal changes/events found (including history): {len(all_rows)}")
    print(f"Going-forward changes (effective_date >= {today}): {len(future_rows)}")

    csv_path = 'index_changes_going_forward.csv'
    write_csv(future_rows, csv_path)
    print(f"\nGoing-forward results written to: {csv_path}")

    xlsx_path = 'index_changes_going_forward.xlsx'
    write_xlsx(future_rows, xlsx_path)
    print(f"Excel report written to: {xlsx_path} "
          f"(sheet 'Side by side' = Removed/Added paired, sheet 'Add remove version' = flat list "
          f"with latest SEC filing for every row, added AND removed)")
    if future_rows:
        for r in future_rows:
            cik_note = f"CIK {r.get('cik')}" if r.get('cik') else "CIK n/a"
            filing_note = format_latest_filing_text(r)
            print(f"  {r.get('effective_date')}  {r.get('action'):8s} {r.get('index_bucket'):16s} "
                  f"{(r.get('ticker') or ''):8s} {r.get('company_name') or r.get('title') or ''} "
                  f"[{cik_note}, {filing_note}]")
    else:
        print("  (none found -- either no upcoming changes were published yet, or a source "
              "failed to fetch; check the log lines above for HTTP status codes and link counts)")

    print()
    email_sent = send_email_report(all_rows, future_rows, csv_path, xlsx_path, today=today)
    if email_sent:
        print(f"Email sent to {SMTP_RECEIVER_EMAIL} from {SMTP_SENDER_EMAIL}.")
    else:
        print(f"Email NOT sent -- see the [Email] warning/error above "
              f"(most likely SMTP_PASSWORD isn't set yet; this needs a Gmail App Password for "
              f"{SMTP_SENDER_EMAIL}, not its regular password).")
